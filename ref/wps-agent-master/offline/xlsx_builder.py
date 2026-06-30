# -*- coding: utf-8 -*-
"""Offline spreadsheet builder & analyzer — openpyxl + pandas.
Requires: pip install openpyxl pandas
"""
from typing import Optional, Dict, List
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


def build_xlsx(structure: Dict, output_path: str) -> Dict:
    """Build .xlsx from JSON structure.
    structure = {
        "sheets": {
            "Sheet1": {
                "column_widths": {"A": 15, "B": 20},
                "rows": [
                    [{"text": "Header1", "bold": True, "bg_color": "4472C4", "font_color": "FFFFFF"}, ...],
                    ["Data1", "Data2"],
                    ["=SUM(B2:B10)"],
                ]
            }
        }
    }
    """
    if not HAS_OPENPYXL:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}
    wb = Workbook()
    first = True
    stats = {"sheets": 0, "rows": 0}
    for sheet_name, sheet_data in structure.get("sheets", {}).items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for col_letter, width in sheet_data.get("column_widths", {}).items():
            ws.column_dimensions[col_letter].width = width
        for r_idx, row_data in enumerate(sheet_data.get("rows", []), 1):
            for c_idx, cell_data in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if isinstance(cell_data, dict):
                    cell.value = cell_data.get("text", "")
                    if cell_data.get("bold"):
                        cell.font = Font(bold=True)
                    if cell_data.get("font_color"):
                        cell.font = Font(color=cell_data["font_color"])
                    if cell_data.get("bg_color"):
                        cell.fill = PatternFill("solid", fgColor=cell_data["bg_color"])
                    if cell_data.get("number_format"):
                        cell.number_format = cell_data["number_format"]
                else:
                    cell.value = cell_data
            stats["rows"] += 1
        stats["sheets"] += 1
    wb.save(output_path)
    return {"output": output_path, "stats": stats}


def analyze_xlsx(filepath: str) -> Dict:
    """Read and analyze .xlsx with pandas: shape, dtypes, head, describe, nulls."""
    if not HAS_PANDAS:
        return {"error": "pandas not installed. Run: pip install pandas"}
    import pandas as pd
    result = {}
    try:
        all_sheets = pd.read_excel(filepath, sheet_name=None)
        for name, df in all_sheets.items():
            result[name] = {
                "shape": list(df.shape),
                "columns": list(df.columns),
                "dtypes": {k: str(v) for k, v in df.dtypes.items()},
                "head": df.head(5).to_dict(orient="records"),
                "describe": df.describe(include="all").to_dict() if any(df.select_dtypes("number").columns) else None,
                "null_count": int(df.isnull().sum().sum()),
                "formulas": _detect_formulas(filepath, name),
            }
        return {"sheets": result, "total_sheets": len(result)}
    except Exception as e:
        return {"error": str(e)}


def _detect_formulas(filepath: str, sheet_name: str) -> int:
    """Count formula cells in a sheet using openpyxl."""
    if not HAS_OPENPYXL:
        return 0
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        ws = wb[sheet_name]
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    count += 1
        wb.close()
        return count
    except Exception:
        return 0


def convert_csv_to_xlsx(csv_path: str, xlsx_path: str, delimiter: str = ",") -> Dict:
    """Convert CSV to XLSX."""
    if not HAS_PANDAS:
        return {"error": "pandas not installed"}
    import pandas as pd
    try:
        df = pd.read_csv(csv_path, delimiter=delimiter)
        df.to_excel(xlsx_path, index=False)
        return {"output": xlsx_path, "rows": len(df), "columns": len(df.columns)}
    except Exception as e:
        return {"error": str(e)}


def validate_formulas_offline(filepath: str) -> Dict:
    """Scan all formula cells in an xlsx for errors (#REF!, #DIV/0!, etc.).
    Uses openpyxl data_only=True to read calculated error values."""
    if not HAS_OPENPYXL:
        return {"error": "openpyxl not installed"}
    ERRORS = ["#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"]
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        result = {}
        total_errors = 0
        for name in wb.sheetnames:
            ws = wb[name]
            sheet_errors = []
            for row in ws.iter_rows():
                for cell in row:
                    val = str(cell.value) if cell.value is not None else ""
                    for err in ERRORS:
                        if err in val:
                            sheet_errors.append({"cell": cell.coordinate, "error": err, "value": val})
                            total_errors += 1
                            break
            if sheet_errors:
                result[name] = sheet_errors
        wb.close()
        return {"status": "errors_found" if total_errors else "success", "total_errors": total_errors, "error_summary": result}
    except Exception as e:
        return {"error": str(e)}


def apply_financial_colors_offline(filepath: str) -> Dict:
    """Apply financial model color conventions to an xlsx file offline."""
    if not HAS_OPENPYXL:
        return {"error": "openpyxl not installed"}
    wb = openpyxl.load_workbook(filepath)
    stats = {"hardcoded": 0, "formula": 0, "internal_ref": 0, "external_ref": 0}
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
                    if ".xls" in formula.lower():
                        cell.font = Font(color="FF0000")
                        stats["external_ref"] += 1
                    elif "!" in formula:
                        cell.font = Font(color="008000")
                        stats["internal_ref"] += 1
                    else:
                        cell.font = Font(color="000000")
                        stats["formula"] += 1
                elif cell.value is not None:
                    cell.font = Font(color="0000FF")
                    stats["hardcoded"] += 1
    wb.save(filepath)
    return {"status": "applied", "stats": stats}
