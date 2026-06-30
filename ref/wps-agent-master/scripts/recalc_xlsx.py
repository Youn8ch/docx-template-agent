# -*- coding: utf-8 -*-
"""Recalculate Excel formulas using LibreOffice headless mode.
Usage: python scripts/recalc_xlsx.py <file.xlsx> [timeout_seconds]
"""
import json
import os
import subprocess
import sys
import tempfile
from pathlib import Path

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def find_soffice() -> str:
    """Find LibreOffice executable."""
    candidates = [
        # Windows
        "C:\\Program Files\\LibreOffice\\program\\soffice.exe",
        "C:\\Program Files (x86)\\LibreOffice\\program\\soffice.exe",
        # Linux
        "/usr/bin/soffice",
        "/usr/lib/libreoffice/program/soffice",
        # macOS
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    # Try PATH
    for cmd in ["soffice", "libreoffice"]:
        try:
            result = subprocess.run([cmd, "--version"], capture_output=True, timeout=5)
            if result.returncode == 0:
                return cmd
        except Exception:
            continue
    return None


def recalc_xlsx(filepath: str, timeout: int = 60) -> dict:
    """Recalculate formulas in .xlsx using LibreOffice and scan for errors."""
    if not os.path.exists(filepath):
        return {"error": f"File not found: {filepath}"}
    if not filepath.lower().endswith((".xlsx", ".xlsm")):
        return {"error": "Not an .xlsx/.xlsm file"}

    soffice = find_soffice()
    if not soffice:
        return {"error": "LibreOffice not found. Install LibreOffice or check PATH.", "note": "Formulas were NOT recalculated. Values may be stale."}

    abs_path = os.path.abspath(filepath)
    work_dir = tempfile.mkdtemp(prefix="recalc_")

    try:
        result = subprocess.run(
            [soffice, "--headless", "--calc", "--norestore",
             f"--convert-to", "xlsx", "--outdir", work_dir, abs_path],
            capture_output=True, text=True, timeout=timeout
        )
    except subprocess.TimeoutExpired:
        return {"error": f"Recalculation timed out after {timeout}s"}

    # Find output file
    base = Path(filepath).stem
    output_files = list(Path(work_dir).glob(f"{base}*.xlsx"))
    if not output_files:
        output_files = list(Path(work_dir).glob("*.xlsx"))

    if output_files:
        output = str(output_files[0])
        os.replace(output, filepath)

    # Scan for formula errors
    if HAS_OPENPYXL:
        errors = scan_formula_errors(filepath)
    else:
        errors = {"note": "openpyxl not installed — cannot scan for formula errors"}

    return {
        "status": "recalculated" if output_files else "no_output",
        "file": filepath,
        "formula_check": errors,
    }


def scan_formula_errors(filepath: str) -> dict:
    """Scan all cells for formula errors using openpyxl."""
    if not HAS_OPENPYXL:
        return {"error": "openpyxl not installed"}
    ERRORS = ["#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"]
    try:
        wb = load_workbook(filepath, data_only=True)
        total_formulas = 0
        errors_found = {}
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    val = str(cell.value) if cell.value is not None else ""
                    for err in ERRORS:
                        if err in val:
                            addr = cell.coordinate
                            if err not in errors_found:
                                errors_found[err] = {"count": 0, "locations": []}
                            errors_found[err]["count"] += 1
                            errors_found[err]["locations"].append(f"{name}!{addr}")
                            break
        wb.close()
        total_errors = sum(v["count"] for v in errors_found.values())
        return {
            "status": "errors_found" if total_errors else "success",
            "total_errors": total_errors,
            "error_summary": errors_found if errors_found else None,
        }
    except Exception as e:
        return {"error": str(e)}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python recalc_xlsx.py <file.xlsx> [timeout_seconds]")
        sys.exit(1)
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 60
    result = recalc_xlsx(sys.argv[1], timeout)
    print(json.dumps(result, indent=2, ensure_ascii=False))
